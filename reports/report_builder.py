from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import matplotlib.pyplot as plt
except Exception:  # pragma: no cover - optional dependency
    plt = None  # type: ignore
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BOOKING_LINK = "https://cal.eu/andrea.ruggiero/sessione-di-confronto-strategico"

BENCHMARK_RANGES = {
    "cogs_pct": (0.25, 0.35),
    "labor_pct": (0.22, 0.35),
    "ebitda_pct": (0.08, 0.18),
}

SCENARIO_ASSUMPTIONS = {
    "conservative": {
        "revenue_mult": 0.90,
        "cogs_pp": 0.02,
        "labor_pp": 0.01,
        "opex_pp": 0.005,
        "marketing_pp": 0.002,
        "fee_pp": 0.0,
    },
    "base": {
        "revenue_mult": 1.00,
        "cogs_pp": 0.0,
        "labor_pp": 0.0,
        "opex_pp": 0.0,
        "marketing_pp": 0.0,
        "fee_pp": 0.0,
    },
    "optimistic": {
        "revenue_mult": 1.08,
        "cogs_pp": -0.01,
        "labor_pp": -0.005,
        "opex_pp": -0.002,
        "marketing_pp": 0.0,
        "fee_pp": 0.0,
    },
}


def _fit_within(src_w: float, src_h: float, max_w: float, max_h: float) -> Tuple[float, float]:
    if src_w <= 0 or src_h <= 0:
        return max_w, max_h
    scale = min(max_w / src_w, max_h / src_h)
    return src_w * scale, src_h * scale


def _logo_size(logo_path: str, max_w: float, max_h: float) -> Tuple[float, float]:
    try:
        w, h = ImageReader(logo_path).getSize()
        return _fit_within(float(w), float(h), float(max_w), float(max_h))
    except Exception:
        return max_w, max_h


def _pdf_page_bg(canvas_obj, doc_obj) -> None:
    w, h = A4
    canvas_obj.saveState()
    canvas_obj.setFillColor(colors.white)
    canvas_obj.rect(0, 0, w, h, stroke=0, fill=1)
    canvas_obj.restoreState()


def _n(v: Any, default: float = 0.0) -> float:
    try:
        if v is None:
            return float(default)
        return float(v)
    except Exception:
        return float(default)


def _fmt_eur(v: Any) -> str:
    x = _n(v, float("nan"))
    if x != x:  # NaN
        return "n/a"
    return f"{x:,.0f} €".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_pct(v: Any) -> str:
    if v is None:
        return "n/a"
    return f"{_n(v) * 100:.1f}%"


def _fmt_num(v: Any) -> str:
    if v is None:
        return "n/a"
    return f"{_n(v):,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_int(v: Any) -> str:
    if v is None:
        return "n/a"
    return f"{_n(v):,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _safe_div(a: Any, b: Any, default: Optional[float] = None) -> Optional[float]:
    num = _n(a, 0.0)
    den = _n(b, 0.0)
    if den == 0:
        return default
    return num / den


def _lang(inputs: Optional[Dict[str, Any]] = None, explicit_lang: Optional[str] = None) -> str:
    if explicit_lang:
        s = str(explicit_lang).upper().strip()
        return "EN" if s == "EN" else "IT"
    if inputs:
        s = str(inputs.get("language", "IT")).upper().strip()
        return "EN" if s == "EN" else "IT"
    return "IT"


def _tr(lang: str, key: str) -> str:
    it = {
        "title": "Economic Model Report",
        "subtitle": "Simulazione economico-finanziaria preliminare per progetti food retail e ristorazione",
        "summary": "Executive summary",
        "input_summary": "Sintesi input",
        "pnl": "Conto economico (P&L) run-rate",
        "be_analysis": "Analisi del break-even",
        "scenario_analysis": "Analisi di scenario",
        "ops_kpis": "KPI operativi",
        "benchmarks": "Benchmark indicativi di mercato",
        "sustainability": "Lettura preliminare di sostenibilità",
        "reality_check": "Reality check / nota metodologica",
        "final_cta": "Vuoi rivedere il progetto con maggiore profondità?",
        "final_cta_body": "Puoi prenotare una call di revisione di 30 minuti per discutere il modello economico, la struttura del break-even e le ipotesi operative principali del progetto.",
        "y1": "Vista Y1 (stagionalita + avviamento)",
        "invest": "Investimenti e ritorni",
        "assessment": "Valutazione finale",
        "notes": "Note narrative",
        "line_item": "Voce",
        "value": "Valore",
        "margin": "% su ricavi",
        "revenue": "Ricavi annui",
        "cogs": "COGS",
        "labor": "Personale",
        "prime_cost": "Prime Cost (COGS + Personale)",
        "opex": "OPEX",
        "marketing": "Marketing",
        "fee": "Fee",
        "occupancy": "Occupancy",
        "ebitda": "EBITDA",
        "be_rev": "Break-even (ricavi annui)",
        "be_orders": "Break-even (ordini/giorno)",
        "cash_inv": "Capitale investito",
        "roi_run": "ROI annuo (run-rate)",
        "payback_run": "Payback (mesi, run-rate)",
        "roi_y1": "ROI annuo (Y1)",
        "payback_y1": "Payback (mesi, Y1)",
        "status": "Esito",
        "business": "Tipologia locale",
        "open_days": "Giorni apertura/mese",
        "ramp": "Mesi avviamento",
        "generated": "Generato il",
        "kpi": "KPI",
        "description": "Questo report si basa sui valori inseriti nel simulatore e rappresenta una valutazione economica preliminare del progetto.",
        "orders_day": "Ordini medi / giorno",
        "avg_ticket": "Scontrino medio",
        "open_days_year": "Giorni apertura / anno",
        "rev_day": "Ricavi medi / giorno",
        "be_rev_day": "Break-even / giorno",
        "be_chart_note": "Il break-even rappresenta il livello di ricavi annui in cui ricavi e costi totali si incontrano.",
        "scenario": "Scenario",
        "conservative": "Prudente",
        "base": "Base",
        "optimistic": "Ottimistico",
        "customers_day": "Clienti richiesti / giorno",
        "within_range": "in linea con il range indicativo",
        "below_range": "sotto il range indicativo",
        "above_range": "sopra il range indicativo",
        "benchmark_note": "I range riportati sono benchmark generici e non vincolanti, utili solo come riferimento orientativo.",
        "reality_note": "Il simulatore non può valutare automaticamente variabili reali come qualità della location, flussi pedonali, intensità competitiva, rotazione tavoli o organizzazione effettiva del personale.",
        "booking_link": "Link prenotazione",
        "disclaimer": "Disclaimer: le valutazioni presenti hanno esclusivo scopo illustrativo e non costituiscono consulenza finanziaria, legale o base sufficiente per decisioni di investimento.",
    }
    en = {
        "title": "Economic Model Report",
        "subtitle": "Preliminary financial simulation for food retail and restaurant projects",
        "summary": "Executive summary",
        "input_summary": "Input summary",
        "pnl": "Run-rate P&L statement",
        "be_analysis": "Break-even analysis",
        "scenario_analysis": "Scenario analysis",
        "ops_kpis": "Operational KPIs",
        "benchmarks": "Indicative market benchmark ranges",
        "sustainability": "Preliminary sustainability reading",
        "reality_check": "Reality check / methodological note",
        "final_cta": "Would you like to review your project in more depth?",
        "final_cta_body": "You can book a 30-minute project review call to discuss the economic model, break-even structure, and key operational assumptions of your project.",
        "y1": "Y1 view (seasonality + ramp-up)",
        "invest": "Investments and returns",
        "assessment": "Final assessment",
        "notes": "Narrative notes",
        "line_item": "Line item",
        "value": "Value",
        "margin": "% of revenue",
        "revenue": "Annual revenue",
        "cogs": "COGS",
        "labor": "Labor",
        "prime_cost": "Prime Cost (COGS + Labor)",
        "opex": "OPEX",
        "marketing": "Marketing",
        "fee": "Fee",
        "occupancy": "Occupancy",
        "ebitda": "EBITDA",
        "be_rev": "Break-even (annual revenue)",
        "be_orders": "Break-even (orders/day)",
        "cash_inv": "Cash invested",
        "roi_run": "Annual ROI (run-rate)",
        "payback_run": "Payback (months, run-rate)",
        "roi_y1": "Annual ROI (Y1)",
        "payback_y1": "Payback (months, Y1)",
        "status": "Outcome",
        "business": "Business type",
        "open_days": "Open days/month",
        "ramp": "Ramp-up months",
        "generated": "Generated on",
        "kpi": "KPI",
        "description": "This report is based on the values entered in the simulator and represents a preliminary economic assessment of the project.",
        "orders_day": "Avg orders / day",
        "avg_ticket": "Average ticket",
        "open_days_year": "Open days / year",
        "rev_day": "Average revenue / day",
        "be_rev_day": "Break-even / day",
        "be_chart_note": "Break-even represents the annual revenue level where revenue and total costs intersect.",
        "scenario": "Scenario",
        "conservative": "Conservative",
        "base": "Base",
        "optimistic": "Optimistic",
        "customers_day": "Required customers / day",
        "within_range": "within the indicative range",
        "below_range": "below the indicative range",
        "above_range": "above the indicative range",
        "benchmark_note": "Ranges shown are generic, non-binding benchmarks and should be used only as indicative references.",
        "reality_note": "The simulator cannot automatically assess real-world variables such as location quality, pedestrian traffic, competition density, table rotation efficiency, or actual labor organization.",
        "booking_link": "Booking link",
        "disclaimer": "Disclaimer: these estimates are for illustrative purposes only and do not constitute financial or legal advice, nor a sufficient basis for investment decisions.",
    }
    return (en if lang == "EN" else it).get(key, key)


def _be_orders(results: Dict[str, Any]) -> Any:
    v = results.get("break_even_orders_day")
    if v is None:
        v = results.get("break_even_orders_per_day")
    return v


def _status_and_messages(feasibility: Optional[Dict[str, Any]]) -> Tuple[str, List[str]]:
    if not isinstance(feasibility, dict):
        return "REVIEW", []
    raw = str(feasibility.get("status") or feasibility.get("label") or "REVIEW").upper()
    if raw == "NO_GO":
        raw = "NO GO"
    msgs = feasibility.get("reasons") or feasibility.get("messages") or feasibility.get("notes") or []
    if not isinstance(msgs, list):
        msgs = [str(msgs)]
    return raw, [str(m) for m in msgs if str(m).strip()]


def _narrative_notes(lang: str, results: Dict[str, Any], status: str) -> List[str]:
    rev = _n(results.get("revenue_annual_runrate"))
    ebitda = _n(results.get("ebitda_annual_runrate"))
    ebitda_pct = _n(results.get("ebitda_pct_annual_runrate"))
    prime = _n(results.get("prime_cost_pct"))
    occ = _n(results.get("occupancy_pct"))
    be = results.get("break_even_revenue_annual")

    if lang == "IT":
        notes = [
            f"Il modello stima ricavi annui run-rate pari a {_fmt_eur(rev)} con EBITDA {_fmt_eur(ebitda)} ({_fmt_pct(ebitda_pct)}).",
            f"Prime Cost al {_fmt_pct(prime)} e Occupancy al {_fmt_pct(occ)}.",
            f"Punto di pareggio stimato a {_fmt_eur(be)}.",
            f"Valutazione complessiva: {status}.",
        ]
    else:
        notes = [
            f"The model estimates run-rate annual revenue at {_fmt_eur(rev)} with EBITDA of {_fmt_eur(ebitda)} ({_fmt_pct(ebitda_pct)}).",
            f"Prime Cost at {_fmt_pct(prime)} and Occupancy at {_fmt_pct(occ)}.",
            f"Estimated break-even revenue at {_fmt_eur(be)}.",
            f"Overall assessment: {status}.",
        ]
    return notes


def _input_summary_rows(inputs: Dict[str, Any], lang: str) -> List[Tuple[str, str]]:
    open_days = int(_n(inputs.get("open_days"), 30))
    rows: List[Tuple[str, str]] = [
        (_tr(lang, "business"), str(inputs.get("business_label", "Custom"))),
        (_tr(lang, "open_days"), str(open_days)),
        (_tr(lang, "open_days_year"), str(open_days * 12)),
    ]
    if _n(inputs.get("orders_per_day")) > 0:
        rows.append((_tr(lang, "orders_day"), _fmt_int(inputs.get("orders_per_day"))))
    if _n(inputs.get("avg_ticket")) > 0:
        rows.append((_tr(lang, "avg_ticket"), _fmt_eur(inputs.get("avg_ticket"))))
    if _n(inputs.get("capex")) > 0:
        rows.append(("CAPEX", _fmt_eur(inputs.get("capex"))))
    if _n(inputs.get("deposits")) > 0:
        rows.append(("Deposits" if lang == "EN" else "Depositi", _fmt_eur(inputs.get("deposits"))))
    if bool(inputs.get("seasonality_enabled")):
        rows.append((_tr(lang, "ramp"), _fmt_int(inputs.get("ramp_up_months"))))
    return rows


def _scenario_models(results: Dict[str, Any]) -> Dict[str, Dict[str, float]]:
    rev = _n(results.get("revenue_annual_runrate"))
    occ = _n(results.get("occupancy_annual_runrate"))
    ratios = {
        "cogs": _safe_div(results.get("cogs_annual_runrate"), rev, 0.0) or 0.0,
        "labor": _safe_div(results.get("labor_annual_runrate"), rev, 0.0) or 0.0,
        "opex": _safe_div(results.get("opex_annual_runrate"), rev, 0.0) or 0.0,
        "marketing": _safe_div(results.get("marketing_annual_runrate"), rev, 0.0) or 0.0,
        "fee": _safe_div(results.get("fee_annual_runrate"), rev, 0.0) or 0.0,
    }
    avg_ticket = _n(results.get("avg_order_value"))
    models: Dict[str, Dict[str, float]] = {}
    for key in ("conservative", "base", "optimistic"):
        conf = SCENARIO_ASSUMPTIONS[key]
        scenario_rev = rev * conf["revenue_mult"]
        cogs = scenario_rev * max(0.0, ratios["cogs"] + conf["cogs_pp"])
        labor = scenario_rev * max(0.0, ratios["labor"] + conf["labor_pp"])
        opex = scenario_rev * max(0.0, ratios["opex"] + conf["opex_pp"])
        marketing = scenario_rev * max(0.0, ratios["marketing"] + conf["marketing_pp"])
        fee = scenario_rev * max(0.0, ratios["fee"] + conf["fee_pp"])
        ebitda = scenario_rev - cogs - labor - opex - marketing - fee - occ
        variable_rate = _safe_div(cogs + labor + opex + marketing + fee, scenario_rev, 0.0) or 0.0
        be_rev = occ / (1 - variable_rate) if scenario_rev > 0 and variable_rate < 1 else 0.0
        models[key] = {
            "revenue": scenario_rev,
            "cogs": cogs,
            "labor": labor,
            "prime_cost": cogs + labor,
            "opex": opex,
            "marketing": marketing,
            "fee": fee,
            "occupancy": occ,
            "ebitda": ebitda,
            "ebitda_pct": _safe_div(ebitda, scenario_rev, 0.0) or 0.0,
            "break_even_revenue": be_rev,
            "break_even_orders": (_safe_div(be_rev / 12.0, avg_ticket, 0.0) or 0.0) if avg_ticket > 0 else 0.0,
        }
    return models


def _scenario_table(results: Dict[str, Any], lang: str) -> List[List[str]]:
    scenario_models = _scenario_models(results)
    labels = {
        "conservative": _tr(lang, "conservative"),
        "base": _tr(lang, "base"),
        "optimistic": _tr(lang, "optimistic"),
    }
    table = [[_tr(lang, "scenario"), _tr(lang, "revenue"), _tr(lang, "ebitda"), _tr(lang, "be_rev")]]
    for key in ("conservative", "base", "optimistic"):
        model = scenario_models[key]
        table.append([
            labels[key],
            _fmt_eur(model["revenue"]),
            _fmt_eur(model["ebitda"]),
            _fmt_eur(model["break_even_revenue"]),
        ])
    return table

def _operational_kpis(inputs: Dict[str, Any], results: Dict[str, Any], lang: str) -> List[Tuple[str, str]]:
    open_days = int(_n(inputs.get("open_days"), 30))
    annual_open_days = max(1, open_days * 12)
    rev = _n(results.get("revenue_annual_runrate"))
    be_rev = _n(results.get("break_even_revenue_annual"))
    be_orders = _be_orders(results)
    rows = [
        (_tr(lang, "revenue"), _fmt_eur(rev)),
        (_tr(lang, "rev_day"), _fmt_eur(rev / annual_open_days if annual_open_days else 0.0)),
    ]
    if _n(inputs.get("orders_per_day")) > 0:
        rows.append((_tr(lang, "orders_day"), _fmt_int(inputs.get("orders_per_day"))))
    if _n(inputs.get("avg_ticket")) > 0:
        rows.append((_tr(lang, "avg_ticket"), _fmt_eur(inputs.get("avg_ticket"))))
    if isinstance(be_orders, (int, float)):
        rows.append((_tr(lang, "customers_day"), _fmt_num(be_orders)))
    if be_rev > 0:
        rows.append((_tr(lang, "be_rev_day"), _fmt_eur(be_rev / annual_open_days)))
    return rows


def _benchmark_position(actual: float, lo: float, hi: float, lang: str) -> str:
    if actual < lo:
        return _tr(lang, "below_range")
    if actual > hi:
        return _tr(lang, "above_range")
    return _tr(lang, "within_range")


def _benchmark_rows(results: Dict[str, Any], lang: str) -> List[List[str]]:
    cogs_pct = _safe_div(results.get("cogs_annual_runrate"), results.get("revenue_annual_runrate"), 0.0) or 0.0
    labor_pct = _safe_div(results.get("labor_annual_runrate"), results.get("revenue_annual_runrate"), 0.0) or 0.0
    ebitda_pct = _n(results.get("ebitda_pct_annual_runrate"))
    items = [
        ("Food cost %" if lang == "EN" else "Food cost %", cogs_pct, BENCHMARK_RANGES["cogs_pct"]),
        ("Labor cost %" if lang == "EN" else "Costo lavoro %", labor_pct, BENCHMARK_RANGES["labor_pct"]),
        ("EBITDA margin %" if lang == "EN" else "Margine EBITDA %", ebitda_pct, BENCHMARK_RANGES["ebitda_pct"]),
    ]
    rows = [[_tr(lang, "kpi"), "Actual" if lang == "EN" else "Attuale", "Range", "Lettura" if lang == "IT" else "Reading"]]
    for label, actual, (lo, hi) in items:
        rows.append([label, _fmt_pct(actual), f"{_fmt_pct(lo)} - {_fmt_pct(hi)}", _benchmark_position(actual, lo, hi, lang)])
    return rows


def _preliminary_assessment(results: Dict[str, Any], feasibility: Optional[Dict[str, Any]], lang: str) -> List[str]:
    status, reasons = _status_and_messages(feasibility)
    ebitda_pct = _n(results.get("ebitda_pct_annual_runrate"))
    labor_pct = _safe_div(results.get("labor_annual_runrate"), results.get("revenue_annual_runrate"), 0.0) or 0.0
    be_orders = _n(_be_orders(results))
    msgs: List[str] = []
    if lang == "IT":
        if ebitda_pct >= 0.15:
            msgs.append("Il modello simulato appare economicamente ben bilanciato con le ipotesi correnti.")
        elif ebitda_pct >= 0.08:
            msgs.append("Il modello risulta sostenibile, ma richiede attenzione alla tenuta dei margini operativi.")
        else:
            msgs.append("Il modello mostra una redditività limitata e richiede un affinamento delle leve economiche.")
        if be_orders > 0:
            if be_orders >= 180:
                msgs.append("Il break-even richiede un traffico giornaliero elevato e va letto con prudenza.")
            elif be_orders >= 100:
                msgs.append("Il break-even appare raggiungibile, ma dipende da una buona continuità di domanda.")
        if labor_pct > 0.30:
            msgs.append("La performance del margine appare sensibile alle ipotesi sul costo del lavoro.")
        if status:
            msgs.append(f"Esito sintetico del simulatore: {status}.")
    else:
        if ebitda_pct >= 0.15:
            msgs.append("The simulated model appears economically well balanced under the current assumptions.")
        elif ebitda_pct >= 0.08:
            msgs.append("The model appears sustainable, but margin resilience should be monitored carefully.")
        else:
            msgs.append("The model shows limited profitability and would benefit from sharper economic assumptions.")
        if be_orders > 0:
            if be_orders >= 180:
                msgs.append("Break-even requires strong daily traffic and should be read cautiously.")
            elif be_orders >= 100:
                msgs.append("Break-even appears achievable, but depends on consistent daily demand.")
        if labor_pct > 0.30:
            msgs.append("Margin performance appears sensitive to labor cost assumptions.")
        if status:
            msgs.append(f"Simulator headline outcome: {status}.")
    msgs.extend(reasons[:2])
    return msgs


def _plot_break_even_png(results: Dict[str, Any], lang: str) -> Optional[bytes]:
    if plt is None:
        return None
    curve = results.get("break_even_curve")
    if not isinstance(curve, dict):
        return None

    xs = curve.get("revenue_annual", [])
    tc = curve.get("total_costs_annual", [])
    eb = curve.get("ebitda_annual", [])
    bex = curve.get("break_even_revenue_annual")
    if not xs or not tc or not eb:
        return None

    fig = plt.figure(figsize=(7.6, 4.8))
    ax = fig.add_subplot(111)
    ax.plot(xs, xs, color="#84665B", linewidth=2.2, label=_tr(lang, "revenue"))
    ax.plot(xs, tc, color="#B89581", linewidth=2.0, linestyle="--", label="Total costs" if lang == "EN" else "Costi totali")
    ax.plot(xs, eb, color="#1C1C1C", linewidth=2.2, label="EBITDA")
    ax.axhline(0, linewidth=0.8, color="#7A7A7A")
    if isinstance(bex, (int, float)) and bex > 0:
        ax.axvline(float(bex), color="#7A7A7A", linestyle=":", linewidth=1.5)
    ax.set_title("Break-even curve" if lang == "EN" else "Curva di break-even")
    ax.set_xlabel("€")
    ax.set_ylabel("€")
    ax.grid(alpha=0.15)
    ax.legend(loc="best")
    fig.tight_layout()
    bio = BytesIO()
    fig.savefig(bio, format="png", dpi=170)
    plt.close(fig)
    return bio.getvalue()


def _plot_pnl_png(results: Dict[str, Any], lang: str) -> Optional[bytes]:
    if plt is None:
        return None
    labels_it = ["Ricavi", "COGS", "Personale", "OPEX", "Marketing", "Fee", "Occupancy", "EBITDA"]
    labels_en = ["Revenue", "COGS", "Labor", "OPEX", "Marketing", "Fee", "Occupancy", "EBITDA"]
    labels = labels_en if lang == "EN" else labels_it
    vals = [
        _n(results.get("revenue_annual_runrate")),
        -_n(results.get("cogs_annual_runrate")),
        -_n(results.get("labor_annual_runrate")),
        -_n(results.get("opex_annual_runrate")),
        -_n(results.get("marketing_annual_runrate")),
        -_n(results.get("fee_annual_runrate")),
        -_n(results.get("occupancy_annual_runrate")),
        _n(results.get("ebitda_annual_runrate")),
    ]
    fig = plt.figure(figsize=(7.6, 4.8))
    ax = fig.add_subplot(111)
    colors_bar = ["#2E2E2E"] + ["#A64A4A"] * 6 + ["#1C1C1C"]
    ax.bar(labels, vals, color=colors_bar)
    ax.set_title("P&L bridge (run-rate)" if lang == "EN" else "Ponte P&L (run-rate)")
    ax.set_ylabel("€")
    ax.grid(axis="y", alpha=0.14)
    plt.xticks(rotation=15)
    fig.tight_layout()
    bio = BytesIO()
    fig.savefig(bio, format="png", dpi=170)
    plt.close(fig)
    return bio.getvalue()


def _plot_cost_mix_png(results: Dict[str, Any], lang: str) -> Optional[bytes]:
    if plt is None:
        return None
    rows = [
        ("COGS", _n(results.get("cogs_annual_runrate"))),
        ("Labor" if lang == "EN" else "Personale", _n(results.get("labor_annual_runrate"))),
        ("OPEX", _n(results.get("opex_annual_runrate"))),
        ("Marketing", _n(results.get("marketing_annual_runrate"))),
        ("Fee", _n(results.get("fee_annual_runrate"))),
        ("Occupancy", _n(results.get("occupancy_annual_runrate"))),
    ]
    rows = [(l, v) for (l, v) in rows if v > 0]
    if not rows:
        return None

    labels = [r[0] for r in rows]
    values = [r[1] for r in rows]
    colors_pie = ["#84665B", "#B89581", "#2E2E2E", "#A67F6F", "#8B817C", "#A64A4A"][: len(values)]
    fig = plt.figure(figsize=(7.6, 6.2))
    ax = fig.add_subplot(111)
    ax.pie(
        values,
        labels=labels,
        autopct="%1.1f%%",
        startangle=90,
        colors=colors_pie,
        wedgeprops={"linewidth": 1, "edgecolor": "white"},
        pctdistance=0.7,
        labeldistance=1.12,
    )
    ax.set_title("Cost mix (annual)" if lang == "EN" else "Mix costi (annuo)")
    fig.tight_layout()
    bio = BytesIO()
    fig.savefig(bio, format="png", dpi=170)
    plt.close(fig)
    return bio.getvalue()


def _plot_daypart_png(inputs: Dict[str, Any], lang: str) -> Optional[bytes]:
    if plt is None:
        return None
    rows = inputs.get("daypart_breakdown") or []
    if not isinstance(rows, list):
        return None

    labels: List[str] = []
    values: List[float] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        label = str(r.get("label", "") or "").strip()
        val = _n(r.get("monthly_revenue"))
        if label and val > 0:
            labels.append(label)
            values.append(val)
    if not labels:
        return None

    fig = plt.figure(figsize=(7.8, 5.8))
    ax = fig.add_subplot(111)
    ax.bar(labels, values, color="#84665B")
    ax.set_title("Daypart breakdown (monthly run-rate)" if lang == "EN" else "Breakdown per fascia (run-rate mensile)")
    ax.set_ylabel("€")
    ax.grid(axis="y", alpha=0.14)
    plt.xticks(rotation=10)
    fig.tight_layout()
    bio = BytesIO()
    fig.savefig(bio, format="png", dpi=170)
    plt.close(fig)
    return bio.getvalue()


def _pnl_rows(results: Dict[str, Any], y1: bool = False) -> List[Tuple[str, float, float]]:
    rev = _n(results.get("revenue_annual_y1")) if y1 else _n(results.get("revenue_annual_runrate"))
    cogs = _n(results.get("cogs_annual_y1")) if y1 else _n(results.get("cogs_annual_runrate"))
    labor = _n(results.get("labor_annual_y1")) if y1 else _n(results.get("labor_annual_runrate"))
    opex = _n(results.get("opex_annual_y1")) if y1 else _n(results.get("opex_annual_runrate"))
    mkt = _n(results.get("marketing_annual_y1")) if y1 else _n(results.get("marketing_annual_runrate"))
    fee = _n(results.get("fee_annual_y1")) if y1 else _n(results.get("fee_annual_runrate"))
    occ = _n(results.get("occupancy_annual_y1")) if y1 else _n(results.get("occupancy_annual_runrate"))
    ebitda = _n(results.get("ebitda_annual_y1")) if y1 else _n(results.get("ebitda_annual_runrate"))
    prime = cogs + labor

    def m(v: float) -> float:
        return (v / rev) if rev > 0 else 0.0

    return [
        ("revenue", rev, 1.0 if rev > 0 else 0.0),
        ("cogs", -cogs, -m(cogs)),
        ("labor", -labor, -m(labor)),
        ("prime_cost", -prime, -m(prime)),
        ("opex", -opex, -m(opex)),
        ("marketing", -mkt, -m(mkt)),
        ("fee", -fee, -m(fee)),
        ("occupancy", -occ, -m(occ)),
        ("ebitda", ebitda, m(ebitda)),
    ]


def build_excel_report_bytes(
    *,
    inputs: Optional[Dict[str, Any]] = None,
    results: Optional[Dict[str, Any]] = None,
    feasibility: Optional[Dict[str, Any]] = None,
    lang: Optional[str] = None,
    logo_path: str = "assets/logo.png",
) -> bytes:
    inputs = inputs or {}
    results = results or {}
    lang = _lang(inputs, lang)
    status, reasons = _status_and_messages(feasibility)
    scenario_models = _scenario_models(results)
    scenario_labels = {
        "conservative": _tr(lang, "conservative"),
        "base": _tr(lang, "base"),
        "optimistic": _tr(lang, "optimistic"),
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Modello" if lang == "IT" else "Model"

    for idx, w in enumerate([34, 22, 22, 22, 18], start=1):
        ws.column_dimensions[chr(64 + idx)].width = w

    thin = Side(style="thin", color="E6E8EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_font = Font(size=12, bold=True, color="1F2937")
    k_font = Font(size=10, bold=True, color="374151")
    v_font = Font(size=11, color="111827")
    subtle_font = Font(size=9, color="6B7280")
    fill_head = PatternFill("solid", fgColor="F3F4F6")
    fill_band = PatternFill("solid", fgColor="F5EFEA")
    fill_highlight = PatternFill("solid", fgColor="FCFCFD")

    for c in ("A", "B", "C", "D", "E"):
        ws[f"{c}1"].fill = fill_band
        ws[f"{c}2"].fill = fill_band
    ws.merge_cells("A1:E2")
    ws["A1"] = ""
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    lp = Path(logo_path)
    if lp.exists():
        try:
            img = XLImage(str(lp))
            lw, lh = _logo_size(str(lp), max_w=420, max_h=110)
            img.width = int(lw)
            img.height = int(lh)
            ws.add_image(img, "A1")
        except Exception:
            pass

    ws.merge_cells("A4:E4")
    ws["A4"] = _tr(lang, "title")
    ws["A4"].font = Font(size=18, bold=True, color="1F2937")
    ws.merge_cells("A5:E5")
    ws["A5"] = f"{_tr(lang, 'subtitle')} - {_tr(lang, 'generated')} {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A5"].font = subtle_font

    row = 7
    ws[f"A{row}"] = _tr(lang, "scenario_analysis")
    ws[f"A{row}"].font = h_font
    row += 1
    ws[f"A{row}"] = _tr(lang, "kpi")
    for col, key in zip(("B", "C", "D"), ("conservative", "base", "optimistic")):
        ws[f"{col}{row}"] = scenario_labels[key]
    for col in ("A", "B", "C", "D"):
        ws[f"{col}{row}"].font = k_font
        ws[f"{col}{row}"].fill = fill_head
        ws[f"{col}{row}"].border = border
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
    row += 1
    for label, metric in (
        (_tr(lang, "revenue"), "revenue"),
        (_tr(lang, "ebitda"), "ebitda"),
        ("EBITDA %", "ebitda_pct"),
        (_tr(lang, "be_rev"), "break_even_revenue"),
    ):
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = k_font
        ws[f"A{row}"].border = border
        for col, key in zip(("B", "C", "D"), ("conservative", "base", "optimistic")):
            value = scenario_models[key][metric]
            ws[f"{col}{row}"] = value
            ws[f"{col}{row}"].border = border
            ws[f"{col}{row}"].alignment = Alignment(horizontal="right")
            ws[f"{col}{row}"].font = v_font
            ws[f"{col}{row}"].number_format = "0.0%" if metric == "ebitda_pct" else '#,##0 [$€-1];[Red]-#,##0 [$€-1]'
        row += 1
    row += 1

    ws[f"A{row}"] = _tr(lang, "pnl")
    ws[f"A{row}"].font = h_font
    row += 1
    ws[f"A{row}"] = _tr(lang, "line_item")
    for col, key in zip(("B", "C", "D"), ("conservative", "base", "optimistic")):
        ws[f"{col}{row}"] = scenario_labels[key]
    ws[f"E{row}"] = _tr(lang, "margin")
    for col in ("A", "B", "C", "D", "E"):
        ws[f"{col}{row}"].font = k_font
        ws[f"{col}{row}"].fill = fill_head
        ws[f"{col}{row}"].border = border
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
    row += 1

    pnl_line_order = ["revenue", "cogs", "labor", "prime_cost", "opex", "marketing", "fee", "occupancy", "ebitda"]
    for metric in pnl_line_order:
        ws[f"A{row}"] = _tr(lang, metric)
        ws[f"A{row}"].font = k_font
        ws[f"A{row}"].border = border
        for col, key in zip(("B", "C", "D"), ("conservative", "base", "optimistic")):
            value = scenario_models[key][metric]
            if metric not in ("revenue", "ebitda") and value > 0:
                value = -value
            ws[f"{col}{row}"] = value
            ws[f"{col}{row}"].number_format = '#,##0 [$€-1];[Red]-#,##0 [$€-1]'
            ws[f"{col}{row}"].font = v_font
            ws[f"{col}{row}"].border = border
            ws[f"{col}{row}"].alignment = Alignment(horizontal="right")
        base_margin = scenario_models["base"]["ebitda_pct"] if metric == "ebitda" else (_safe_div(scenario_models["base"][metric], scenario_models["base"]["revenue"], 0.0) or 0.0)
        if metric not in ("revenue", "ebitda"):
            base_margin *= -1
        ws[f"E{row}"] = 1.0 if metric == "revenue" else base_margin
        ws[f"E{row}"].number_format = "0.0%"
        ws[f"E{row}"].font = v_font
        ws[f"E{row}"].border = border
        ws[f"E{row}"].alignment = Alignment(horizontal="right")
        if metric in ("revenue", "ebitda"):
            for col in ("A", "B", "C", "D", "E"):
                ws[f"{col}{row}"].fill = fill_highlight
        row += 1
    row += 1

    ws[f"A{row}"] = _tr(lang, "ops_kpis")
    ws[f"A{row}"].font = h_font
    row += 1
    ws[f"A{row}"] = _tr(lang, "kpi")
    ws[f"B{row}"] = _tr(lang, "value")
    for col in ("A", "B"):
        ws[f"{col}{row}"].font = k_font
        ws[f"{col}{row}"].fill = fill_head
        ws[f"{col}{row}"].border = border
    row += 1
    for label, value in _operational_kpis(inputs, results, lang):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = k_font
        ws[f"B{row}"].font = v_font
        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border
        ws[f"B{row}"].alignment = Alignment(horizontal="right")
        row += 1
    row += 1

    ws[f"A{row}"] = _tr(lang, "assessment")
    ws[f"A{row}"].font = h_font
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws[f"A{row}"] = f"{_tr(lang, 'status')}: {status}"
    ws[f"A{row}"].font = Font(size=11, bold=True, color="1F2937")
    row += 1
    for r in reasons:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws[f"A{row}"] = f"- {r}"
        ws[f"A{row}"].font = subtle_font
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws[f"A{row}"] = f"{_tr(lang, 'booking_link')}: {BOOKING_LINK}"
    ws[f"A{row}"].font = Font(size=9, bold=True, color="84665B")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws[f"A{row}"] = _tr(lang, 'disclaimer')
    ws[f"A{row}"].font = Font(size=8, italic=True, color="6B7280")
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A8"

    ws2 = wb.create_sheet("Grafici" if lang == "IT" else "Charts")
    ws2.column_dimensions["A"].width = 2
    ws2.column_dimensions["B"].width = 44
    ws2.column_dimensions["C"].width = 44
    ws2["B2"] = "Grafici di supporto" if lang == "IT" else "Supporting charts"
    ws2["B2"].font = h_font
    try:
        be_png = _plot_break_even_png(results, lang)
    except Exception:
        be_png = None
    try:
        pnl_png = _plot_pnl_png(results, lang)
    except Exception:
        pnl_png = None
    try:
        cost_png = _plot_cost_mix_png(results, lang)
    except Exception:
        cost_png = None
    try:
        daypart_png = _plot_daypart_png(inputs, lang)
    except Exception:
        daypart_png = None
    if be_png:
        img1 = XLImage(BytesIO(be_png))
        img1.width = 640
        img1.height = 320
        ws2.add_image(img1, "B4")
    if pnl_png:
        img2 = XLImage(BytesIO(pnl_png))
        img2.width = 640
        img2.height = 320
        ws2.add_image(img2, "B25")
    if cost_png:
        img3 = XLImage(BytesIO(cost_png))
        img3.width = 600
        img3.height = 360
        ws2.add_image(img3, "B46")
    if daypart_png:
        img4 = XLImage(BytesIO(daypart_png))
        img4.width = 640
        img4.height = 320
        ws2.add_image(img4, "B69")

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def build_pdf_report_bytes(
    *,
    inputs: Optional[Dict[str, Any]] = None,
    results: Optional[Dict[str, Any]] = None,
    feasibility: Optional[Dict[str, Any]] = None,
    lang: Optional[str] = None,
    logo_path: str = "assets/logo.png",
) -> bytes:
    inputs = inputs or {}
    results = results or {}
    lang = _lang(inputs, lang)
    status, reasons = _status_and_messages(feasibility)
    notes = _narrative_notes(lang, results, status)
    input_rows = _input_summary_rows(inputs, lang)
    scenario_rows = _scenario_table(results, lang)
    ops_rows = _operational_kpis(inputs, results, lang)
    benchmark_rows = _benchmark_rows(results, lang)
    sustainability_notes = _preliminary_assessment(results, feasibility, lang)

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.4 * cm,
        bottomMargin=1.4 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title_sp", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=16, leading=20, textColor=colors.HexColor("#1F2937"))
    h_style = ParagraphStyle("h_sp", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12, leading=15, spaceBefore=8, textColor=colors.HexColor("#1F2937"))
    body = ParagraphStyle("body_sp", parent=styles["BodyText"], fontName="Helvetica", fontSize=9.5, leading=13, textColor=colors.HexColor("#374151"))
    small = ParagraphStyle("small_sp", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.8, leading=12, textColor=colors.HexColor("#4B5563"))
    cta_style = ParagraphStyle("cta_sp", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=11, leading=15, textColor=colors.HexColor("#1F2937"))
    cta_button_style = ParagraphStyle("cta_button_sp", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=10.5, leading=13, alignment=1, textColor=colors.white)

    story: List[Any] = []
    lp = Path(logo_path)
    if lp.exists():
        try:
            lw, lh = _logo_size(str(lp), max_w=9.4 * cm, max_h=3.0 * cm)
            story.append(Image(str(lp), width=lw, height=lh))
        except Exception:
            pass
    story.append(Spacer(1, 0.14 * cm))
    story.append(Paragraph(_tr(lang, "title"), title_style))
    story.append(Paragraph(f"{_tr(lang, 'subtitle')} - {_tr(lang, 'generated')} {datetime.now().strftime('%Y-%m-%d %H:%M')}", small))
    story.append(Spacer(1, 0.34 * cm))

    story.append(Paragraph(_tr(lang, "summary"), h_style))
    story.append(Paragraph(_tr(lang, "description"), body))
    story.append(Spacer(1, 0.16 * cm))
    exec_data = [
        [_tr(lang, "kpi"), _tr(lang, "value"), _tr(lang, "kpi"), _tr(lang, "value")],
        [_tr(lang, "business"), str(inputs.get("business_label", "Custom")), _tr(lang, "open_days"), str(int(_n(inputs.get("open_days"), 30)))],
        [_tr(lang, "revenue"), _fmt_eur(results.get("revenue_annual_runrate")), _tr(lang, "ebitda"), _fmt_eur(results.get("ebitda_annual_runrate"))],
        ["EBITDA %", _fmt_pct(results.get("ebitda_pct_annual_runrate")), _tr(lang, "be_rev"), _fmt_eur(results.get("break_even_revenue_annual"))],
        [_tr(lang, "be_orders"), _fmt_num(_be_orders(results)), _tr(lang, "cash_inv"), _fmt_eur(results.get("cash_invested"))],
        [_tr(lang, "roi_run"), _fmt_pct(results.get("roi_annual")), _tr(lang, "payback_run"), _fmt_num(results.get("payback_months"))],
    ]
    t_exec = Table(exec_data, colWidths=[4.7 * cm, 3.1 * cm, 4.7 * cm, 3.1 * cm])
    t_exec.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFCFD")]),
    ]))
    story.append(t_exec)
    story.append(Spacer(1, 0.22 * cm))

    story.append(Paragraph(_tr(lang, "input_summary"), h_style))
    input_data = [[_tr(lang, "line_item"), _tr(lang, "value")]] + [[label, value] for label, value in input_rows]
    t_input = Table(input_data, colWidths=[8.6 * cm, 5.8 * cm])
    t_input.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFCFD")]),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
    ]))
    story.append(t_input)
    story.append(Spacer(1, 0.16 * cm))

    story.append(Paragraph(_tr(lang, "pnl"), h_style))
    pnl_data = [[_tr(lang, "line_item"), _tr(lang, "value"), _tr(lang, "margin")]]
    for k, v, p in _pnl_rows(results, y1=False):
        pnl_data.append([_tr(lang, k), _fmt_eur(v), _fmt_pct(p)])
    t_pnl = Table(pnl_data, colWidths=[7.8 * cm, 3.6 * cm, 3.2 * cm])
    t_pnl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFCFD")]),
    ]))
    story.append(t_pnl)
    story.append(Spacer(1, 0.16 * cm))

    story.append(Paragraph(_tr(lang, "y1"), h_style))
    y1_data = [[_tr(lang, "line_item"), _tr(lang, "value"), _tr(lang, "margin")]]
    for k, v, p in _pnl_rows(results, y1=True):
        y1_data.append([_tr(lang, k), _fmt_eur(v), _fmt_pct(p)])
    t_y1 = Table(y1_data, colWidths=[7.8 * cm, 3.6 * cm, 3.2 * cm])
    t_y1.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFCFD")]),
    ]))
    story.append(t_y1)
    story.append(Spacer(1, 0.16 * cm))

    try:
        be_png = _plot_break_even_png(results, lang)
    except Exception:
        be_png = None
    story.append(Paragraph(_tr(lang, "be_analysis"), h_style))
    be_data = [
        [_tr(lang, "line_item"), _tr(lang, "value")],
        [_tr(lang, "be_rev"), _fmt_eur(results.get("break_even_revenue_annual"))],
        [_tr(lang, "be_orders"), _fmt_num(_be_orders(results))],
        (_tr(lang, "be_rev_day"), _fmt_eur((_n(results.get("break_even_revenue_annual")) / max(1, int(_n(inputs.get("open_days"), 30)) * 12)))),
    ]
    t_be = Table(be_data, colWidths=[8.6 * cm, 5.8 * cm])
    t_be.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFCFD")]),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
    ]))
    story.append(t_be)
    story.append(Spacer(1, 0.08 * cm))
    story.append(Paragraph(_tr(lang, "be_chart_note"), small))
    if be_png:
        story.append(Spacer(1, 0.08 * cm))
        story.append(Image(BytesIO(be_png), width=17.0 * cm, height=7.9 * cm))
    story.append(Spacer(1, 0.16 * cm))

    story.append(Paragraph(_tr(lang, "scenario_analysis"), h_style))
    t_scen = Table(scenario_rows, colWidths=[3.6 * cm, 4.3 * cm, 4.3 * cm, 4.4 * cm])
    t_scen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFCFD")]),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
    ]))
    story.append(t_scen)
    story.append(Spacer(1, 0.16 * cm))

    story.append(Paragraph(_tr(lang, "ops_kpis"), h_style))
    ops_data = [[_tr(lang, "kpi"), _tr(lang, "value")]] + [[label, value] for label, value in ops_rows]
    t_ops = Table(ops_data, colWidths=[8.6 * cm, 5.8 * cm])
    t_ops.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFCFD")]),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
    ]))
    story.append(t_ops)
    story.append(Spacer(1, 0.16 * cm))

    story.append(Paragraph(_tr(lang, "invest"), h_style))
    inv_data = [
        [_tr(lang, "line_item"), _tr(lang, "value")],
        [_tr(lang, "cash_inv"), _fmt_eur(results.get("cash_invested"))],
        [_tr(lang, "roi_run"), _fmt_pct(results.get("roi_annual"))],
        [_tr(lang, "payback_run"), _fmt_num(results.get("payback_months"))],
        [_tr(lang, "roi_y1"), _fmt_pct(results.get("roi_annual_y1"))],
        [_tr(lang, "payback_y1"), _fmt_num(results.get("payback_months_y1"))],
    ]
    t_inv = Table(inv_data, colWidths=[7.8 * cm, 6.8 * cm])
    t_inv.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFCFD")]),
    ]))
    story.append(t_inv)
    story.append(Spacer(1, 0.16 * cm))

    story.append(Paragraph(_tr(lang, "benchmarks"), h_style))
    t_bench = Table(benchmark_rows, colWidths=[4.7 * cm, 2.6 * cm, 3.6 * cm, 4.9 * cm])
    t_bench.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFCFD")]),
        ("ALIGN", (1, 1), (2, -1), "RIGHT"),
    ]))
    story.append(t_bench)
    story.append(Spacer(1, 0.08 * cm))
    story.append(Paragraph(_tr(lang, "benchmark_note"), small))
    story.append(Spacer(1, 0.16 * cm))

    story.append(Paragraph(_tr(lang, "sustainability"), h_style))
    for item in sustainability_notes:
        story.append(Paragraph(f"- {item}", body))
    story.append(Spacer(1, 0.1 * cm))

    if notes:
        story.append(Paragraph(_tr(lang, "notes"), h_style))
        for item in notes:
            story.append(Paragraph(f"- {item}", body))
        story.append(Spacer(1, 0.1 * cm))

    story.append(Paragraph(_tr(lang, "reality_check"), h_style))
    story.append(Paragraph(_tr(lang, "reality_note"), body))
    story.append(Spacer(1, 0.12 * cm))

    try:
        pnl_png = _plot_pnl_png(results, lang)
    except Exception:
        pnl_png = None
    try:
        cost_png = _plot_cost_mix_png(results, lang)
    except Exception:
        cost_png = None
    try:
        daypart_png = _plot_daypart_png(inputs, lang)
    except Exception:
        daypart_png = None
    if reasons:
        story.append(Paragraph(_tr(lang, "assessment"), h_style))
        story.append(Paragraph(f"<b>{_tr(lang, 'status')}:</b> {status}", body))
        for r in reasons:
            story.append(Paragraph(f"- {r}", body))
        story.append(Spacer(1, 0.12 * cm))

    if pnl_png or cost_png or daypart_png:
        story.append(PageBreak())
        story.append(Paragraph("Grafici di supporto" if lang == "IT" else "Supporting charts", h_style))
        story.append(Spacer(1, 0.08 * cm))
    if pnl_png:
        story.append(Table([[Image(BytesIO(pnl_png), width=15.8 * cm, height=6.6 * cm)]], colWidths=[17.8 * cm], style=TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ])))
        story.append(Spacer(1, 0.14 * cm))
    if cost_png:
        story.append(Table([[Image(BytesIO(cost_png), width=13.8 * cm, height=7.4 * cm)]], colWidths=[17.8 * cm], style=TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ])))
        story.append(Spacer(1, 0.14 * cm))
    if daypart_png:
        story.append(Table([[Image(BytesIO(daypart_png), width=15.8 * cm, height=6.6 * cm)]], colWidths=[17.8 * cm], style=TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ])))
        story.append(Spacer(1, 0.12 * cm))

    story.append(PageBreak())
    story.append(Paragraph(_tr(lang, "final_cta"), h_style))
    story.append(Paragraph(_tr(lang, "final_cta_body"), body))
    story.append(Spacer(1, 0.06 * cm))
    cta_label = "Prenota la call di revisione" if lang == "IT" else "Book the review call"
    cta_button = Table(
        [[Paragraph(f"<link href='{BOOKING_LINK}' color='white'>{cta_label}</link>", cta_button_style)]],
        colWidths=[8.6 * cm],
    )
    cta_button.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#111827")),
        ("BOX", (0, 0), (0, 0), 0, colors.white),
        ("VALIGN", (0, 0), (0, 0), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "CENTER"),
        ("LEFTPADDING", (0, 0), (0, 0), 12),
        ("RIGHTPADDING", (0, 0), (0, 0), 12),
        ("TOPPADDING", (0, 0), (0, 0), 10),
        ("BOTTOMPADDING", (0, 0), (0, 0), 10),
        ("ROUNDEDCORNERS", [8, 8, 8, 8]),
    ]))
    story.append(cta_button)
    story.append(Spacer(1, 0.08 * cm))
    story.append(Paragraph(f"{_tr(lang, 'booking_link')}: <font color='#84665B'>{BOOKING_LINK}</font>", cta_style))
    story.append(Spacer(1, 0.18 * cm))
    story.append(Paragraph(f"<i>{_tr(lang, 'disclaimer')}</i>", small))

    doc.build(story, onFirstPage=_pdf_page_bg, onLaterPages=_pdf_page_bg)
    return bio.getvalue()
